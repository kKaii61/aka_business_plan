import base64
from odoo import api, fields, models, _
import io
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook
from odoo.modules.module import get_module_resource
from openpyxl.styles import numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


class ReportNumberOfWorker(models.TransientModel):
    _name = 'report.number.of.worker'
    _rec_name = 'name'

    lookup_date = fields.Date(string='Ngày tra cứu')
    name = fields.Char(string="Tên", default='BÁO CÁO SỐ LƯỢNG LAO ĐỘNG')

    line_ids = fields.One2many('report.number.of.worker.line', 'report_id', string='report line')

    @staticmethod
    def query_data_report():
        query = """
            with worker as (
                select 
                hj.id as job_category_id,
                hj.name as job_category_name,
                
                count(DISTINCT he.id) as worker,
                
                case
                    when he.gender = 'male' then count(distinct he.id)
                end as male,
                
                case 
                    when he.gender = 'female' then count(distinct he.id)
                end as female,
                
                case 
                    when he.certificate in ('doctor', 'associate_professor', 'professor') then count(distinct he.id)
                end as on_doctor_degree,
                case 
                    when he.certificate = 'master' then count(distinct he.id)
                end as master_degree,
                case 
                    when he.certificate = 'bachelor' then count(distinct he.id)
                end as bachelor_degree,
                case 
                    when he.certificate = 'college' then count(distinct he.id)
                end as college_level,
                case 
                    when he.certificate in ('basic_vocational', 'intermediate_vocational') then count(distinct he.id)
                end as intermediate_degree,
                case 
                    when he.certificate in ('general', 'other') then count(distinct he.id)
                end as general_edu_level
                
                from hr_employee he 
                join hr_job hj on hj.id = he.job_id 
                join hr_contract hc on hc.employee_id = he.id
                
                where
                hc.state in ('open')
                {lookup_date}
                
                group by hj.id, hj.name, he.certificate, he.gender
            ),
            
            worker_sum as (
                select
                job_category_id, job_category_name,
                sum(worker) as quantity_worker,
                sum(male) as quantity_male,
                sum(female) as quantity_female,
                sum(on_doctor_degree) as quantity_on_doctor_degree,
                sum(master_degree) as quantity_master_degree,
                sum(bachelor_degree) as quantity_bachelor_degree,
                sum(college_level) as quantity_college_level,
                sum(intermediate_degree) as quantity_intermediate_degree,
                sum(general_edu_level) as quantity_general_edu_level
            
                from worker
                group by job_category_id, job_category_name
            ),
            
            over_12_month as (
                select 
                hj.id as job_category_id,
                hj.name as job_category_name,
                count(DISTINCT he.id) as quantity
                
                from hr_employee he 
                join hr_job hj on hj.id = he.job_id 
                join hr_contract hc on hc.employee_id = he.id
                
                where
                case 
                    when hc.date_end is null then  '1' = '1'
                    else hc.date_start + INTERVAL '1 year' <= hc.date_end
                end
                and hc.state in ('open')
                {lookup_date}
                
                group by hj.id, hj.name
            )
            
            select ws.*,
            om.quantity as quantity_over_12_month
            
            from worker_sum ws
            left join over_12_month om on om.job_category_id = ws.job_category_id
            
            order by job_category_id
        """
        return query

    def get_data(self):
        lookup_date = ""
        if self.lookup_date:
            lookup_date = """
                and cast((hc.date_start + INTERVAL '7 hours') as date) <= '{date}'
            """.format(date=self.lookup_date)

        query = self.query_data_report()
        query = query.format(lookup_date=lookup_date)
        self.env.cr.execute(query)

        res = self.env.cr.dictfetchall()
        return res

    def open_report_view(self):
        try:
            self.line_ids = None
            line_ids = self.get_data()
            values = []
            for line in line_ids:
                vals = {
                    'x_working_position_id': line.get('job_category_id'),
                    'x_job_position': line.get('job_category_name'),
                    'quantity_worker': line.get('quantity_worker'),
                    'quantity_male': line.get('quantity_male'),
                    'quantity_female': line.get('quantity_female'),
                    'quantity_over_12_month': line.get('quantity_over_12_month'),
                    'quantity_on_doctor_degree': line.get('quantity_on_doctor_degree'),
                    'quantity_master_degree': line.get('quantity_master_degree'),
                    'quantity_bachelor_degree': line.get('quantity_bachelor_degree'),
                    'quantity_college_level': line.get('quantity_college_level'),
                    'quantity_intermediate_degree': line.get('quantity_intermediate_degree'),
                    'quantity_general_edu_level': line.get('quantity_general_edu_level'),
                    'report_id': self.id
                }
                values.append((0, 0, vals))

            self.line_ids = values
        except Exception as e:
            raise ValueError(e)

    def write_a_cell(self, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    def format_number_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row)]
            if currentCell.value:
                if currentCell.value % 1 == 0:
                    currentCell.number_format = '#,###'
                else:
                    currentCell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    def add_border_excel(self, ws, index_row, column_quantity):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for column in range(column_quantity):
            ws.cell(row=index_row, column=column + 1).border = thin_border

    def text_align_center_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row + 1)]
            currentCell.alignment = Alignment(horizontal='center')

    def open_report_excel(self):
        filename = 'report_number_of_worker.xlsx'
        existing_filepath = get_module_resource('enmasys_hr_employee', 'static', filename)
        wb = load_workbook(existing_filepath)
        ws = wb.active

        data_to_write = self.get_data()
        start_row = 2
        index_row = 0
        count = 0

        for row in data_to_write:
            count += 1
            line = [
                count,
                row["x_job_position"],
                row["quantity_worker"],
                row["quantity_male"],
                row["quantity_female"],
                row["quantity_over_12_month"],
                row["quantity_on_doctor_degree"],
                row["quantity_master_degree"],
                row["quantity_bachelor_degree"],
                row["quantity_college_level"],
                row["quantity_intermediate_degree"],
                row["quantity_general_edu_level"],
            ]
            self.write_a_cell(ws, line, start_row, index_row)
            index_row += 1
            self.add_border_excel(ws, index_row + 1, column_quantity=12)
            self.text_align_center_excel(ws, index_row, ['A'])

        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.env['report.number.of.worker.excel'].create(
            {'excel_file': base64.encodestring(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/report.number.of.worker.excel/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res


class ReportNumberOfWorkerLine(models.TransientModel):
    _name = 'report.number.of.worker.line'

    report_id = fields.Many2one('report.number.of.worker')

    x_working_position_id = fields.Many2one(comodel_name="hr.job", string="Vị trí công việc")
    x_job_position = fields.Char(string='Vị trí công việc')
    quantity_worker = fields.Integer(string='Số lượng người lao động')
    quantity_male = fields.Integer(string='Nam')
    quantity_female = fields.Integer(string='Nữ')
    quantity_over_12_month = fields.Integer(string='HĐ 12 tháng trở lên')
    quantity_on_doctor_degree = fields.Integer(string='Tiến sĩ')
    quantity_master_degree = fields.Integer(string='Thạc sĩ')
    quantity_bachelor_degree = fields.Integer(string="Cử nhân")
    quantity_college_level = fields.Integer(string="Tốt nghiệp")
    quantity_intermediate_degree = fields.Integer(string='Sơ cấp/ Trung cấp')
    quantity_general_edu_level = fields.Integer(string='LĐPT')


class ReportNumberOfWorkerExcel(models.TransientModel):
    _name = 'report.number.of.worker.excel'

    excel_file = fields.Binary('Report file')
    file_name = fields.Char('Excel file', size=64)
