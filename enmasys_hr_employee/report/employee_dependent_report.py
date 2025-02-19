import logging
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment
import io
import base64

from odoo import models, fields, api, _
from odoo.modules.module import get_module_resource


_logger = logging.getLogger(__name__)


class EmployeeDependentReport(models.TransientModel):
    _name = "employee.dependent.report"
    _description = "Report: Employee Dependents"
    _rec_name = 'name'

    _EMPLOYEE_BLOCKS = {
        'supply': _("Supply"),
        'office': _("Office")
    }

    # filter-fields
    x_department_ids = fields.Many2many(
        comodel_name="hr.department", string="Phòng ban",
        relation="employee_dependent_report_department_rel", column1="report_id", column2="department_id")
    x_employee_ids = fields.Many2many(
        comodel_name="hr.employee", string="Nhân viên",
        relation="employee_dependent_report_employee_rel", column1="report_id", column2="employee_id")
    x_relationship_type_ids = fields.Many2many(
        comodel_name="relatives.type", string="Mối quan hệ",
        relation="employee_dependent_report_relationship_type_rel", column1="report_id", column2="type_id")
    name = fields.Char(string="Tên", default='BÁO CÁO THÔNG TIN NGƯỜI THÂN')

    # data-fields
    x_report_data_ids = fields.One2many(
        comodel_name="employee.dependent.report.data", string="Report datas", inverse_name="x_report_id")

    def name_get(self):
        names = []
        DEFAULT_NAME = _("Report: Employee Dependents")
        for report in self:
            names.append((report.id, DEFAULT_NAME))
        return names


    def _generate_query_filter(
            self, departments=None, employees=None, relationship_types=None,
            current_user=None, current_company=None):
        try:
            where_clause = str()
            if not departments and not employees and not relationship_types:
                return where_clause
            where_clause = "WHERE"
            default_filters = [
                "dcl.x_hr_employee_id IS NOT NULL"
            ]
            filters = default_filters
            if departments:
                filters.append("he.department_id IN %(department_ids)s" % {
                    'department_ids': '(' + ', '.join(map(str, departments.ids)) + ')'
                })
            if employees:
                filters.append("he.id IN %(employee_ids)s" % {
                    'employee_ids': '(' + ', '.join(map(str, employees.ids)) + ')'
                })
            if relationship_types:
                filters.append("dcl.x_dependent_contact_relationship_id IN %(type_ids)s" % {
                    'type_ids': '(' + ', '.join(map(str, relationship_types.ids)) + ')'
                })
            if current_user:
                _company = current_company or current_user.company_id
                _current_employee = self.env['hr.employee'].sudo().search([
                    ('user_id', '=', current_user.id), ('company_id', '=', _company.id)
                ], limit=1)
                if bool(_current_employee and _current_employee.sudo().department_id.x_department_code == 'PVD-MP' and
                        current_user.has_group(group_ext_id='enmasys_hr_employee.group_employee_hr_other') and
                        not current_user.has_group(group_ext_id='enmasys_hr_employee.group_supply_contract_manager')):
                    filters.append(f"he.x_block = \'supply\'")
            where_clause = f"{where_clause} {' AND '.join(filters)}"
            return where_clause
        except Exception as e:
            _logger.exception(msg=e)

    def _generate_report_datas(self):
        QUERY_STATEMENT = """
            SELECT dcl.id AS dependence_contact_id
            FROM hr_employee_dependence_contact_line AS dcl
            INNER JOIN  hr_employee AS he ON he.id = dcl.x_hr_employee_id
            {WHERE};
        """
        # as result, the Report Datas will have format
        # [val_1, val_2, ...] : val_1, val_2, ... will be dictionary
        filter_clause = self._generate_query_filter(
            employees=self.x_employee_ids, departments=self.x_department_ids,
            relationship_types=self.x_relationship_type_ids,
            current_user=self.env.user, current_company=self.env.company)
        self.env.cr.execute(QUERY_STATEMENT.format(WHERE=filter_clause))
        report_datas = self.env.cr.dictfetchall()
        return report_datas

    def generate_report_datas(self):

        self.x_report_data_ids.unlink()
        datas = self._generate_report_datas()
        report_datas = self.env['employee.dependent.report.data'].create([{
            'x_report_id': self.id,
            'x_dependence_line_id': data.get('dependence_contact_id')
        } for data in datas])
        return report_datas


    def make_report_exportation(self):

        # check report-datas
        # if nothing now, too lazy to do anything
        current_datas = self.x_report_data_ids
        if not current_datas:
            return current_datas

        exportation_file_name = _("Employee Dependents Report")
        exportation_template_name = "employee_dependents_report_exportation_template.xlsx"
        exportation_path = get_module_resource('enmasys_hr_employee', 'report', exportation_template_name)
        if not exportation_path:
            return

        exportation_workbook = load_workbook(exportation_path)
        exportation_worksheet = exportation_workbook.active

        # make table-content
        self._make_table_contents(worksheet=exportation_worksheet, report_datas=current_datas)

        # save exportation
        fp = io.BytesIO()
        exportation_workbook.save(fp)

        # create attachment
        new_attachment = self.env['ir.attachment'].sudo().create({
            'name': exportation_file_name,
            'datas': base64.encodebytes(fp.getvalue()),
            'res_model': self._name,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        # download attachment
        return {
            "type": "ir.actions.act_url",
            "name": exportation_file_name,
            "url": '/web/content/' + str(new_attachment.sudo().id) + '?download=true',
            "target": "new",
        }


    @classmethod
    def _make_cell_fit_content(cls, worksheet, cell, width_rate):
        cell_content_width = len(str(cell.value))
        current_column_width = worksheet.column_dimensions[cell.column_letter].width or 10
        new_column_width = max(current_column_width, cell_content_width) * width_rate
        worksheet.column_dimensions[cell.column_letter].width = new_column_width


    @classmethod
    def _get_col_position_and_data_field_collection(cls):

        # format:
        # {column_idx: data_field, ...}
        data_collection = {
            1: {'order': int()},
            2: {'x_employee_ref': str()},  # employee-ref
            3: {'x_employee_name': str()},  # employee-name
            4: {'x_employee_block': str()},  # employee-block
            5: {'x_employee_gender': str()},  # employee-gender
            6: {'x_employee_department_id': str()},  # employee-department
            7: {'x_dependent_name': str()},  # dependent-name
            8: {'x_dependent_dob': str()},  # dependent-dob
            9: {'x_dependent_relationship_type_id': str()},  # relation-type-id
        }
        return data_collection


    def _from_raw_datas_generate_table_contents(self, raw_datas):
        table_contents = list()
        employee_blocks = self._EMPLOYEE_BLOCKS
        employee_genders = dict(self.env['hr.employee']._fields['gender'].selection)
        for raw_data in raw_datas:
            data_collection = self._get_col_position_and_data_field_collection()
            for data_order, _data in data_collection.items():
                _data_field = list(data_collection[data_order])[0]
                _content = str()
                if _data_field == 'order':
                    data_collection[data_order][_data_field] = len(table_contents) + 1
                elif _data_field in ['x_dependent_relationship_type_id', 'x_employee_department_id']:
                    data_collection[data_order][_data_field] = raw_data[_data_field].name or str()
                elif _data_field == 'x_dependent_dob' and raw_data[_data_field]:
                    data_collection[data_order][_data_field] = raw_data[_data_field].strftime('%d/%m/%Y')
                elif _data_field in ['x_employee_block']:
                    data_collection[data_order][_data_field] = _(employee_blocks.get(raw_data[_data_field], str()))
                elif _data_field in ['x_employee_gender']:
                    data_collection[data_order][_data_field] = _(employee_genders.get(raw_data[_data_field], str()))
                else:
                    data_collection[data_order][_data_field] = raw_data[_data_field] or str()
            table_contents.append(data_collection)
        return table_contents


    def _make_table_contents(self, worksheet, report_datas, first_row_position=5):

        # if nothing, too lazy to do anything
        if not report_datas:
            return report_datas
        # init formatted-values
        _SIDE = Side(border_style='thin', color="FF000000")
        _BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
        _ALIGNMENT = Alignment(horizontal='center')
        # try set table-content to worksheet
        table_contents = self._from_raw_datas_generate_table_contents(raw_datas=report_datas)
        for _table_content in table_contents:
            for column_idx, column_content in _table_content.items():
                _content_cell = worksheet.cell(row=first_row_position, column=column_idx)
                _content_cell.value = list(column_content.values())[0]
                _content_cell.border = _BORDER
                if column_idx in [1, 2, 4, 5, 8]:
                    _content_cell.alignment = _ALIGNMENT
                self._make_cell_fit_content(worksheet=worksheet, cell=_content_cell, width_rate=1.0)
            first_row_position += 1

