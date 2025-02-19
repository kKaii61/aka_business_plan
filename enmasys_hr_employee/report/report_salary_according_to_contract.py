import base64
from odoo import api, fields, models, _
import io
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook
from odoo.modules.module import get_module_resource
from openpyxl.styles import numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


class ReportSalaryAccordingToContract(models.TransientModel):
    _name = 'report.salary.according.to.contract'
    _rec_name = 'name'

    department_ids = fields.Many2many('hr.department', string='Phòng ban')

    contract_type_ids = fields.Many2many('hr.contract.type',
                                         'hr_contract_type_report_salary_according_to_contract_rel',
                                         'report_salary_according_to_contract_id', 'hr_contract_type_id',
                                         string='Loại hợp đồng')
    contract_status = fields.Selection([('open', 'Còn hiệu lực'), ('close', 'Hết hiệu lực')],
                                       string='Trạng thái hợp đồng')
    from_salary = fields.Integer(string='Mức lương từ', default=0)
    to_salary = fields.Integer(string='Mức lương đến', default=0)

    employee_ids = fields.Many2many('hr.employee', string='Nhân viên')
    name = fields.Char(string="Tên", default='BÁO CÁO LƯƠNG THEO HỢP ĐỒNG')

    line_ids = fields.One2many('report.salary.according.to.contract.line', 'report_id', string='report line')

    @api.onchange("department_ids")
    def onchange_employee_ids(self):
        domain = [('department_id', 'in', self.department_ids.ids)] if self.department_ids else []
        return {'domain': {'employee_ids': domain}}

    def query_data_report(self):
        query = """
                select 
                    he.name as employee_name,
                    hj.name as job_name,
                    hd.name as department_name,
                    rpb.sanitized_acc_number as bank_account,
                    hct.name as contract_type,
                    hc.wage as salary,
                    hc.wage as allowance_amount_total,
                    
                    hc.employee_id,
                    hc.job_id,
                    hc.department_id
                    
                from hr_contract hc 
                    left JOIN hr_employee he on he.id = hc.employee_id
                    left JOIN hr_job hj on hj.id = hc.job_id
                    left JOIN hr_department hd on hd.id = hc.department_id
                    left JOIN res_partner_bank rpb on rpb.id = he.bank_account_id 
                    left JOIN hr_contract_type hct on hct.id = hc.contract_type_id 
                    
                where 
	                hc.state in ('open', 'close')
                    and (hd.id in ({department}) or '0' = '{department}')
                    and (he.id in ({employee}) or '0' = '{employee}')

                    {contract_status}
                    {from_salary_condition}
                    {to_salary_condition}
        """
        return query

    def get_data(self):
        department_ids = ','.join(
            [str(idd) for idd in self.department_ids.ids]) if self.department_ids else '0'
        contract_type_ids = ','.join(
            [str(idd) for idd in self.contract_type_ids.ids]) if self.contract_type_ids else '0'
        employee_ids = ','.join(
            [str(idd) for idd in self.employee_ids.ids]) if self.employee_ids else '0'

        contract_status = "and hc.state = '%s'" % self.contract_status if self.contract_status else ''

        from_salary_condition = "and hc.wage >= '%s'" % self.from_salary if self.from_salary else ''
        to_salary_condition = "and hc.wage <= '%s'" % self.to_salary if self.to_salary else ''

        query = self.query_data_report()
        query = query.format(department=department_ids, contract_type=contract_type_ids,
                             employee=employee_ids,
                             contract_status=contract_status,
                             from_salary_condition=from_salary_condition, to_salary_condition=to_salary_condition)
        self.env.cr.execute(query)

        res = self.env.cr.dictfetchall()
        return res

    def open_report_view(self):
        try:
            self.line_ids = None
            line_ids = self.get_data()
            values = []
            lang = self.env.user.lang
            for line in line_ids:
                vals = {
                    'employee_id': line.get('employee_id'),
                    'job_id': line.get('job_id'),
                    'department_id': line.get('department_id'),
                    'bank_name': line.get('bank_name'),
                    'bank_account': line.get('bank_account'),
                    'contract_type': line.get('contract_type').get(lang) if line.get('contract_type') else '',
                    'salary': line.get('salary'),
                    'allowance_amount_total': line.get('allowance_amount_total'),
                    'report_id': self.id,
                }
                values.append((0, 0, vals))

            self.line_ids = values
        except Exception as e:
            raise ValueError(e)

    def write_a_cell(self, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    def format_number_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row)]
            if currentCell.value:
                if currentCell.value % 1 == 0:
                    currentCell.number_format = '#,###'
                else:
                    currentCell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    def add_border_excel(self, ws, index_row, column_quantity):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for column in range(column_quantity):
            ws.cell(row=index_row, column=column + 1).border = thin_border

    def text_align_center_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row + 1)]
            currentCell.alignment = Alignment(horizontal='center')

    def open_report_excel(self):
        filename = 'report_salary_according_to_contract.xlsx'
        existing_filepath = get_module_resource('enmasys_hr_employee', 'report', filename)
        wb = load_workbook(existing_filepath)
        ws = wb.active

        data_to_write = self.get_data()
        start_row = 2
        index_row = 0
        count = 0
        lang = self.env.user.lang

        for row in data_to_write:
            count += 1
            job_name = row.get('job_name').get(lang) if row.get('job_name') else ''
            department_name = row.get('department_name').get(lang) if row.get('department_name') else ''
            contract_type = row.get('contract_type').get(lang) if row.get('contract_type') else ''
            line = [
                count,
                row["employee_name"],
                job_name,
                department_name,
                row["bank_account"],
                contract_type,
                row["salary"],
                row["allowance_amount_total"],
            ]
            self.write_a_cell(ws, line, start_row, index_row)
            index_row += 1
            self.add_border_excel(ws, index_row + 1, column_quantity=8)
            self.text_align_center_excel(ws, index_row, ['A'])
            self.format_number_excel(ws, index_row + 1, ['M', 'N'])

        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.env['report.salary.according.to.contract.excel'].create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/report.salary.according.to.contract.excel/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res

    def name_get(self):
        try:
            names = list()
            for report in self:
                names.append((report.id, _("Employee-Contract's Salary")))
            return names
        except Exception as e:
            raise ValidationError(e)


class ReportSalaryAccordingToContractLine(models.TransientModel):
    _name = 'report.salary.according.to.contract.line'
    _order = "salary desc"

    report_id = fields.Many2one('report.salary.according.to.contract')

    employee_id = fields.Many2one('hr.employee', string='Họ và tên')
    job_id = fields.Many2one('hr.job', string='Vị trí')
    department_id = fields.Many2one('hr.department', string='Phòng ban')

    bank_name = fields.Char(string='Ngân hàng')
    bank_account = fields.Char(string='Số tài khoản')
    contract_type = fields.Char(string='Loại hợp đồng')
    job_level = fields.Char(string='Bậc')
    salary = fields.Integer(string='Lương')
    allowance_amount_total = fields.Integer(string='Tổng phụ cấp')


class ReportSalaryAccordingToContractExcel(models.TransientModel):
    _name = 'report.salary.according.to.contract.excel'

    excel_file = fields.Binary('Report file')
    file_name = fields.Char('Excel file', size=64)
