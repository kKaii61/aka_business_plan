import base64
from odoo import api, fields, models, _
import io
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook
from odoo.modules.module import get_module_resource
from openpyxl.styles import numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


class ContractEmployeeReport(models.TransientModel):
    _name = 'contract.employee.report'
    _rec_name= 'name'

    x_department_ids = fields.Many2many(comodel_name="hr.department", string='Phòng ban')

    x_contract_date_from = fields.Date(string='Ngày hợp đồng từ')
    x_contract_date_arrive = fields.Date(string='Ngày hợp đồng đến')
    x_end_date_from = fields.Date(string='Ngày kết thúc từ')
    x_end_date_arrive = fields.Date(string='Ngày kết thúc đến')
    x_contract_ids = fields.Many2many(comodel_name="hr.contract", string='Hợp đồng')
    x_staff_ids = fields.Many2many(comodel_name="hr.employee", string='Nhân viên')
    x_contract_status = fields.Selection([
        ('still_validated', 'Còn hiệu lực'),
        ('no_longer_valid', 'Hết hiệu lực')
    ], string="Trạng thái hợp đồng")
    x_information_contract_staff_ids = fields.One2many("contract.employee.report.line",
                                                       "x_information_contract_staff_id")
    name = fields.Char(string="Tên",default='BÁO CÁO HỢP ĐỒNG NHÂN VIÊN')

    def query_data_report(self):
        query = """
            with query as (
                select  
                    he.name as employee_name,
                    hc.employee_id as employee_id,
                    hj.name as job_name,
                    hc.job_id as job_id,
                    hp.name as department_name,
                    case
                        when he.gender = 'male' then 'Nam'
                        when he.gender = 'female' then 'Nữ'
                        when he.gender = 'other' then 'Khác'
                    end as x_gender,
                    hc.department_id as department_id,
                    hc.name as x_contract_code,
                    hc.date_start as x_date_from ,
                    hc.date_end as x_date_arrive,
                    case 
                        when hc.state = 'draft' then 'Mới'
                        when hc.state = 'open' then 'Đang chạy'
                        when hc.state = 'close' then 'Đã hết hạn'
                        when hc.state = 'cancel' then 'Đã hũy'
                    end as x_status_contract,
                    case 
                        when hc.state = 'draft' then 'still_validated'
                        when hc.state = 'open' then 'still_validated'
                        when hc.state = 'close' then 'no_longer_valid'
                        when hc.state = 'cancel' then 'no_longer_valid'
                    end as status,
                    hct.name as hct_name
                                    
                from hr_contract hc
                    left join hr_employee he on he.id = hc.employee_id
                    left join hr_job hj on hj.id = hc.job_id
                    left join hr_department hp on hp.id = hc.department_id
                    left join hr_contract_type hct on hct.id = hc.contract_type_id
                    
                    
                where
                hc.state in ('open', 'close')
                and hc.employee_id is not null
                and (hc.department_id in ({department}) or '0' = '{department}')
                and (hc.id in ({contract}) or '0' = '{contract}')
                and (hc.employee_id in ({staff}) or '0' = '{staff}')
                {x_contract_date_from}
                {x_contract_date_arrive}
                {x_end_date_from}
                {x_end_date_arrive}
            )
            
            select * from query
            where
                '1'='1'
                {status}
        """
        return query

    def get_data(self):
        x_department_ids = ','.join(
            [str(idd) for idd in self.x_department_ids.ids]) if self.x_department_ids else '0'
        x_contract_ids = ','.join(
            [str(idd) for idd in self.x_contract_ids.ids]) if self.x_contract_ids else '0'
        x_staff_ids = ','.join(
            [str(idd) for idd in self.x_staff_ids.ids]) if self.x_staff_ids else '0'
        x_contract_date_from = "and cast((hc.date_start + INTERVAL '7 hours') as date) >= '%s'" % self.x_contract_date_from if self.x_contract_date_from else ''
        x_contract_date_arrive = "and cast((hc.date_start + INTERVAL '7 hours') as date) <= '%s'" % self.x_contract_date_arrive if self.x_contract_date_arrive else ''
        x_end_date_from = "and cast((hc.date_end + INTERVAL '7 hours') as date) >= '%s'" % self.x_end_date_from if self.x_end_date_from else ''
        x_end_date_arrive = "and cast((hc.date_end + INTERVAL '7 hours') as date) <= '%s'" % self.x_end_date_arrive if self.x_end_date_arrive else ''
        status = "and status = '%s'" % self.x_contract_status if self.x_contract_status else ""

        query = self.query_data_report()
        query = query.format(department=x_department_ids,
                             # contract_type=x_contract_type_ids,
                             contract=x_contract_ids, staff=x_staff_ids,
                             x_contract_date_from=x_contract_date_from,
                             x_contract_date_arrive=x_contract_date_arrive,
                             x_end_date_from=x_end_date_from,
                             x_end_date_arrive=x_end_date_arrive,
                             status=status)
        self.env.cr.execute(query)
        res = self.env.cr.dictfetchall()
        return res

    def open_report_view(self):
        try:
            self.x_information_contract_staff_ids = None
            x_information_contract_staff_ids = self.get_data()
            lang = self.env.user.lang
            values = []
            for line in x_information_contract_staff_ids:
                vals = {
                    'employee_id': line.get('employee_id'),
                    'job_id': line.get('job_id'),
                    'x_gender': line.get('x_gender'),
                    'department_id': line.get('department_id'),
                    'x_contract_code': line.get('x_contract_code'),
                    'x_type_contract': line.get('hct_name').get(lang) if line.get('hct_name') else '',
                    'x_date_from': line.get('x_date_from'),
                    'x_date_arrive': line.get('x_date_arrive'),
                    'x_status_contract': line.get('x_status_contract'),
                    # 'x_information_contract_staff_id': self.id
                }
                values.append((0, 0, vals))

            self.x_information_contract_staff_ids = values
        except Exception as e:
            raise ValueError(e)

    @staticmethod
    def write_a_cell(ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    @staticmethod
    def add_border_excel(ws, index_row, column_quantity):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for column in range(column_quantity):
            ws.cell(row=index_row, column=column + 1).border = thin_border

    def text_align_center_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row + 1)]
            currentCell.alignment = Alignment(horizontal='center')

    def open_report_excel(self):
        pass
        filename = 'report_contract_employee.xlsx'
        existing_filepath = get_module_resource('enmasys_hr_employee', 'report', filename)
        wb = load_workbook(existing_filepath)
        ws = wb.active

        data_to_write = self.get_data()
        start_row = 2
        index_row = 0
        count = 0
        lang = self.env.user.lang
        for row in data_to_write:
            hct_name = row.get('hct_name').get(lang) if row.get('hct_name') else ''
            job_name = row.get('job_name').get(lang) if row.get('job_name') else ''
            department_name = row.get('department_name').get(lang) if row.get('department_name') else ''
            count += 1
            line = [
                count,
                '',
                row["employee_name"],
                job_name,
                row["x_gender"],
                department_name,
                row["x_contract_code"],
                hct_name,
                row["x_date_from"].strftime('%d/%m/%Y') if row["x_date_from"] else '',
                row["x_date_arrive"].strftime('%d/%m/%Y') if row["x_date_arrive"] else '',
                row["x_status_contract"],
            ]
            self.write_a_cell(ws, line, start_row, index_row)
            index_row += 1
            # self.format_number_excel(ws, index_row + 1, ['K'])
            self.add_border_excel(ws, index_row + 1, column_quantity=11)
            self.text_align_center_excel(ws, index_row, ['A'])

        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.env['contract.employee.report.excel'].create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/contract.employee.report.excel/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res


class ContractEmployeeReportLine(models.TransientModel):
    _name = 'contract.employee.report.line'

    x_information_contract_staff_id = fields.Many2one(comodel_name="contract.employee.report")
    x_staff_code = fields.Char(string='Mã nhân viên')
    employee_id = fields.Many2one(comodel_name="hr.employee", string='Tên nhân viên')
    job_id = fields.Many2one('hr.job', string='Vị trí làm việc')
    x_gender = fields.Char(string='Giới tính')
    department_id = fields.Many2one('hr.department', string='Phòng ban')
    x_contract_code = fields.Char(string='Mã hợp đồng')
    x_type_contract = fields.Char(string='Loại hợp đồng')
    x_date_from = fields.Date(string='Từ ngày')
    x_date_arrive = fields.Date(string='Đến ngày')
    x_status_contract = fields.Char(string='Trạng thái')


class ContractEmployeeReportExcel(models.TransientModel):
    _name = 'contract.employee.report.excel'

    excel_file = fields.Binary('Report file')
    file_name = fields.Char('Excel file', size=64)
