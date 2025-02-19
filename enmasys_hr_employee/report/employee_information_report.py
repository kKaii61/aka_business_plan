from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
from odoo.modules.module import get_module_resource
from openpyxl import load_workbook
import io
import base64


class EmployeeInformationReport(models.TransientModel):
    _name = 'employee.information.report'
    _rec_name = 'name'

    x_department_ids = fields.Many2many('hr.department', string='Phòng ban')
    x_gender = fields.Selection(
        [('male', 'Nam'), ('female', 'Nữ'), ('other', 'Khác')],
        string="Giới tính")
    x_marital = fields.Selection([
        ('single', 'Độc thân'),
        ('married', 'Đã cuới'),
        ('cohabitant', 'Legal Cohabitant'),
        ('widower', 'Widower'),
        ('divorced', 'Divorced')
    ], string='Hôn nhân')
    name = fields.Char(string="Tên", default='BÁO CÁO THÔNG TIN NHÂN VIÊN')

    x_job_ids = fields.Many2many('hr.job', string='Chức vụ')
    # x_first_contract_date = fields.Date('Ngày vào làm việc')
    # x_union = fields.x_marital = fields.Selection(
    #     [('doan_vien', 'Đoàn viên'), ('dang_vien', 'Đảng viên'), ('thanh_nien', 'Thanh niên')],
    #     string="Đoàn/ Đảng")
    # month_greater_than = fields.Integer('Số tháng làm việc/ năm từ')
    # month_less_than = fields.Integer('Số tháng làm việc/năm đến')
    x_certificate = fields.Selection([
        ('general', "Phổ thông"),
        ('basic_vocational', "Sơ cấp nghề"),
        ('intermediate_vocational', "Trung cấp nghề"),
        ('college', "Cao đẳng"),
        ('bachelor', 'Đại học'),
        ('master', 'Thạc sĩ'),
        ('doctor', 'Tiến sĩ'),
        ('associate_professor', 'Phó giáo sư'),
        ('professor', 'Giáo sư'),
        ('other', 'Khác'),
    ], string="Trình độ")
    x_employee_information_report_line_ids = fields.One2many('employee.information.report.line',
                                                             'x_employee_information_report', string='Report line ids')

    def get_data(self):
        self.x_employee_information_report_line_ids = None
        department_ids = '(' + ','.join([str(idd) for idd in self.x_department_ids.ids]) + ')'
        job_ids = '(' + ','.join([str(idd) for idd in self.x_job_ids.ids]) + ')'
        job_id_condition = ('and job_id in %s' % job_ids) if self.x_job_ids else ''
        gender_condition = ("and gender = '%s'" % self.x_gender) if self.x_gender else ''
        marital_condition = ("and marital = '%s'" % self.x_marital) if self.x_marital else ''
        department_id_condition = ('and department_id in %s' % department_ids) if self.x_department_ids else ''
        certificate_condition = ("and certificate = '%s'" % self.x_certificate) if self.x_certificate else ''
        query = """
            select  distinct he.id, (CURRENT_DATE - he.first_contract_date) as day_of_service,
                    he.name as employee_name,
                    he.job_id as employee_position,
                    he.gender as employee_gender,
                    he.department_id as employee_department,
                    he.certificate as employee_certificate,
                  
                    he.marital as employee_marital
                   
            from hr_employee as he
            where
                '1' = '1' 
                %s
                %s
                %s
                %s
                %s

        """
        # current_user = self.env.user
        # if (current_user.x_hr_employee_id
        #         and current_user.x_hr_employee_id.department_id
        #         and current_user.x_hr_employee_id.department_id.x_department_code == 'PVD-MP'
        #         and current_user.has_group('enmasys_hr_employee.group_employee_hr_other')):
        #     employee_block = f" AND he.x_block = \'supply\'"
        # else:
        #     employee_block = str()
        self.env.cr.execute(query % (
            department_id_condition, gender_condition, marital_condition, job_id_condition, certificate_condition))
        data = self.env.cr.dictfetchall()
        values = []
        for line in data:
            vals = {
                # 'x_block': line.get('employee_block'),
                'x_employee_code': line.get('employee_code'),
                'x_fullname': line.get('employee_name'),
                'x_job_id': line.get('employee_position'),
                'x_gender': line.get('employee_gender'),
                'x_department_id': line.get('employee_department'),
                'x_certificate': line.get('employee_certificate'),
                # 'x_years_of_service': round(line.get('day_of_service') / 365) if line.get('day_of_service') else 0,
                # 'x_union': line.get('employee_x_position'),
                # 'x_union_in': line.get('employee_date_in'),
                # 'x_first_contract_date': line.get('employee_first_contract_date'),
                'x_marital': line.get('employee_marital')
            }
            values.append((0, 0, vals))
        self.x_employee_information_report_line_ids = values
        return data

    def write_a_cell(self, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    def open_report_excel(self):
        filename = 'report_employee_information.xlsx'
        existing_filepath = get_module_resource('enmasys_hr_employee', 'report', filename)
        wb = load_workbook(existing_filepath)
        ws = wb.active

        data_to_write = self.get_data()
        start_row = 4
        index_row = 0
        for row in data_to_write:
            line = [
                '',
                row["employee_name"],
                # row['employee_block'],
                self.env['hr.job'].browse(row["employee_position"]).name if self.env['hr.job'].browse(row["employee_position"]) else '',
                row["employee_gender"],
                self.env['hr.department'].browse(row["employee_department"]).name if self.env['hr.department'].browse(
                    row["employee_department"]) else '',
                row["employee_certificate"],
                # row["day_of_service"],
                # row["employee_x_position"],
                # row["employee_date_in"],
                # row["employee_first_contract_date"],
                row["employee_marital"],
            ]
            self.write_a_cell(ws, line, start_row, index_row)
            index_row += 1

        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.env['employee.information.excel'].create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/employee.information.excel/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res


class EmployeeInformationReportLine(models.TransientModel):
    _name = 'employee.information.report.line'

    x_employee_code = fields.Char('Mã nhân viên')
    x_fullname = fields.Char('Họ và tên')
    x_job_id = fields.Many2one('hr.job', 'Vị trí')
    x_gender = fields.Selection(
        [('male', 'Nam'), ('female', 'Nữ'), ('other', 'Khác')],
        string="Giới tính")
    x_department_id = fields.Many2one('hr.department', string='Phòng ban')
    x_certificate = fields.Selection([
        ('general', "Phổ thông"),
        ('basic_vocational', "Sơ cấp nghề"),
        ('intermediate_vocational', "Trung cấp nghề"),
        ('college', "Cao đẳng"),
        ('bachelor', 'Đại học'),
        ('master', 'Thạc sĩ'),
        ('doctor', 'Tiến sĩ'),
        ('associate_professor', 'Phó giáo sư'),
        ('professor', 'Giáo sư'),
        ('other', 'Khác'),
    ], string="Trình độ")
    # x_union = fields.x_marital = fields.Selection(
    #     [('doan_vien', 'Đoàn viên'), ('dang_vien', 'Đảng viên'), ('thanh_nien', 'Thanh niên')],
    #     string="Đoàn/ Đảng")
    # x_union_in = fields.Date('Ngày vào Đoàn/ Đảng')
    # x_first_contract_date = fields.Date('Ngày vào làm việc')
    x_marital = fields.Selection([
        ('single', 'Độc thân'),
        ('married', 'Đã cuới'),
        ('cohabitant', 'Legal Cohabitant'),
        ('widower', 'Widower'),
        ('divorced', 'Divorced')
    ], string='Hôn nhân')
    x_employee_information_report = fields.Many2one('employee.information.report', 'Report id')
    # x_years_of_service = fields.Integer('Số năm làm việc')
    # x_block = fields.Char(string="Employee's block")


class EmployeeInformationExcel(models.TransientModel):
    _name = 'employee.information.excel'

    excel_file = fields.Binary('Report file')
    file_name = fields.Char('Excel file', size=64)
