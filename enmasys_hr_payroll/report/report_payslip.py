import base64
from datetime import datetime

from odoo import api, fields, models, _
import io
from openpyxl import load_workbook
from odoo.modules.module import get_module_resource
from openpyxl.styles import numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from odoo.exceptions import UserError, ValidationError


class ReportPayslip(models.TransientModel):
    _name = 'report.payslip'
    _rec_name = 'name'

    from_month = fields.Selection(
        [('1', '1'), ('2', '2'), ('3', '3'), ('4', '4'), ('5', '5'), ('6', '6'), ('7', '7'), ('8', '8'), ('9', '9'),
         ('10', '10'), ('11', '11'), ('12', '12')],
        string='Từ tháng', required=True)
    to_month = fields.Selection(
        [('1', '1'), ('2', '2'), ('3', '3'), ('4', '4'), ('5', '5'), ('6', '6'), ('7', '7'), ('8', '8'), ('9', '9'),
         ('10', '10'), ('11', '11'), ('12', '12')],
        string='Đến tháng', required=True)
    year = fields.Char(string='Năm', default=datetime.now().year, required=True)
    department_ids = fields.Many2many('hr.department', string='Phòng ban')
    name = fields.Char(string="Tên", default='BÁO CÁO PHIẾU LƯƠNG')

    line_ids = fields.One2many('report.payslip.line', 'report_id', string='report line')

    def query_data_report(self):
        query = """
            with allowance as (
                select 
                
                he.name as employee_name,
                hj.name as job_name,
                hd.name as department_name,
                hp.number as payslip_code,
                hp.name as payslip_name,
	           
                
                sum(hpl.total) as total_allowance,
                
                he.id as employee_id,
                hj.id as job_id,
                hd.id as department_id,
                hp.id as payslip_id
                
                from hr_payslip hp
                left join hr_employee he on he.id = hp.employee_id 
                left join hr_job hj on hj.id = he.job_id
                left join hr_department hd on hd.id = he.department_id
                left join hr_contract hc on hc.id = hp.contract_id 
                left join hr_payslip_line hpl on hpl.slip_id = hp.id
                left join hr_salary_rule_category hsrc on hsrc.id = hpl.category_id
                
                where
                    hsrc.code = 'PC'
                    and hp.state in ('verify', 'done')
		            and hc.state in ('open', 'close')
                    and (hd.id in ({department}) or '0' = '{department}')
                    and extract(year from hp.date_from) = {year}
                    and extract(month from hp.date_from) >= {from_month}
                    and extract(month from hp.date_to) <= {to_month}
                
                group by  employee_name, job_name, department_name, 
                payslip_code, payslip_name, he.id, hj.id, hd.id, hp.id
            ),
            
            gross as (
                select 
                
                he.name as employee_name,
                hj.name as job_name,
                hd.name as department_name,
                hp.number as payslip_code,
                hp.name as payslip_name,
	         
                
                sum(hpl.total) as total_salary,
                
                he.id as employee_id,
                hj.id as job_id,
                hd.id as department_id,
                hp.id as payslip_id
                
                from hr_payslip hp
                left join hr_employee he on he.id = hp.employee_id 
                left join hr_job hj on hj.id = he.job_id
                left join hr_department hd on hd.id = he.department_id
                left join hr_contract hc on hc.id = hp.contract_id 
                left join hr_payslip_line hpl on hpl.slip_id = hp.id
                
                where
                    hpl.code = 'GROSS'
		            and hp.state in ('verify', 'done')
		            and hc.state in ('open', 'close')
                    and (hd.id in ({department}) or '0' = '{department}')
                    and extract(year from hp.date_from) = {year}
                    and extract(month from hp.date_from) >= {from_month}
                    and extract(month from hp.date_to) <= {to_month}
                
                group by  employee_name, job_name, department_name, 
                payslip_code, payslip_name,  he.id, hj.id, hd.id, hp.id
            )
            
            select g.*, a.total_allowance
            from gross g
            left join allowance a on  g.employee_id = a.employee_id
                              and g.payslip_id = a.payslip_id
                             
        """
        return query

    def get_data(self):
        department_ids = ','.join(
            [str(idd) for idd in self.department_ids.ids]) if self.department_ids else '0'

        year = int(self.year)
        from_month = self.from_month
        to_month = self.to_month

        query = self.query_data_report()
        query = query.format(department=department_ids, year=year,
                             from_month=from_month, to_month=to_month)
        self.env.cr.execute(query)

        res = self.env.cr.dictfetchall()
        return res

    def check_filter_month(self):
        if int(self.from_month) > int(self.to_month):
            raise UserError(_("Từ tháng phải nhỏ hơn hoặc bằng Đến tháng"))

    def open_report_view(self):
        self.check_filter_month()
        try:
            self.line_ids = None
            line_ids = self.get_data()
            values = []
            for line in line_ids:
                vals = {
                    # 'employee_code': line.get('employee_code'),
                    'employee_id': line.get('employee_id'),
                    'job_id': line.get('job_id'),
                    'department_id': line.get('department_id'),
                    'payslip_code': line.get('payslip_code'),
                    'payslip_id': line.get('payslip_id'),
                    'payslip_name': line.get('payslip_name'),

                    'total_allowance': line.get('total_allowance'),
                    'total_salary': line.get('total_salary'),
                    'report_id': self.id
                }
                values.append((0, 0, vals))

            self.line_ids = values
        except Exception as e:
            raise ValueError(e)

    def write_a_cell(self, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    def format_number_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row)]
            if currentCell.value:
                if currentCell.value % 1 == 0:
                    currentCell.number_format = '#,###'
                else:
                    currentCell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    def add_border_excel(self, ws, index_row, column_quantity):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for column in range(column_quantity):
            ws.cell(row=index_row, column=column + 1).border = thin_border

    def text_align_center_excel(self, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row + 1)]
            currentCell.alignment = Alignment(horizontal='center')

    def open_report_excel(self):
        self.check_filter_month()
        filename = 'report_payslip.xlsx'
        existing_filepath = get_module_resource('enmasys_hr_payroll', 'report', filename)
        wb = load_workbook(existing_filepath)
        ws = wb.active

        data_to_write = self.get_data()
        start_row = 2
        index_row = 0
        count = 0

        for row in data_to_write:
            count += 1
            line = [
                count,
                # row["employee_code"],
                row["employee_name"],
                row["job_name"]['vi_VN'] if row["job_name"] else None,
                row["department_name"]['vi_VN'] if row["department_name"] else None,
                row["payslip_code"],
                row["payslip_name"],
                # row["basic_salary"],
                row["total_allowance"],
                row["total_salary"],
            ]
            self.write_a_cell(ws, line, start_row, index_row)
            index_row += 1
            self.add_border_excel(ws, index_row + 1, column_quantity=11)
            self.text_align_center_excel(ws, index_row, ['A'])
            self.format_number_excel(ws, index_row + 1, ['H', 'I', 'J'])

        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.env['report.payslip.excel'].create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/report.payslip.excel/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res


class ReportPayslipLine(models.TransientModel):
    _name = 'report.payslip.line'

    report_id = fields.Many2one('report.payslip')

    employee_code = fields.Char(string='Mã nhân viên')
    employee_id = fields.Many2one('hr.employee', string='Họ và tên')
    job_id = fields.Many2one('hr.job', string='Vị trí')
    department_id = fields.Many2one('hr.department', string='Phòng ban')

    payslip_code = fields.Char(string='Mã phiếu')
    payslip_id = fields.Many2one('hr.payslip', string='Phiếu lương')
    payslip_name = fields.Char(string='Tên phiếu lương')
    basic_salary = fields.Integer(string='Lương cơ bản')
    total_allowance = fields.Float(string='Tổng phụ cấp')
    total_salary = fields.Float(string='Tổng lương')


class ReportPayslipExcel(models.TransientModel):
    _name = 'report.payslip.excel'

    excel_file = fields.Binary('Report file')
    file_name = fields.Char('Excel file', size=64)
