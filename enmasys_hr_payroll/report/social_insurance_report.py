import logging
import base64
import io
from openpyxl.styles import Border, Side
from openpyxl import load_workbook

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from odoo.tools.misc import file_path

_logger = logging.getLogger(__name__)


class SocialInsuranceReport(models.TransientModel):
    _name = "social.insurance.report"
    _description = "Social Insurance Report"

    # filters
    x_month_from = fields.Selection(
        selection=[(str(month), str(month) if month >= 10 else f"0{month}") for month in range(1, 13)],
        string="From month", required=True)
    x_month_to = fields.Selection(
        selection=[(str(month), str(month) if month >= 10 else f"0{month}") for month in range(1, 13)],
        string="To month", required=True)
    x_year = fields.Integer(string="Year", required=True)
    display_name = fields.Char(default=_("Social Insurance Report"))

    @api.model
    def default_get(self, default_fields):
        try:
            defaults = super(SocialInsuranceReport, self).default_get(default_fields)
            _its = fields.Date.today()
            defaults['x_month_from'] = defaults['x_month_to'] = str(_its.month)
            defaults['x_year'] = _its.year
            return defaults
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    x_salary_report_ids = fields.One2many(
        comodel_name="social.insurance.salary.report", string="Salaries Report", inverse_name="x_report_id")

    @api.onchange('x_month_from', 'x_month_to')
    def _onchange_month_filters(self):
        try:
            for report in self:
                report.x_month_to = max(report.x_month_from, report.x_month_to)
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    @classmethod
    def _generate_query_filter(cls, month_from=None, month_to=None, year=None):
        try:
            """
            hps: hr_payslip
            """
            where_clause = str()
            if not month_from and not month_to and not year and not year:
                return where_clause
            where_clause = "WHERE"
            default_filters = [
                "hps.state = 'done'",
                "hc.state ='open'",
            ]
            filters = default_filters
            if month_from and month_to:
                filters.append("DATE_PART('MONTH', hps.date_to) BETWEEN %(_from_month)s AND %(_to_month)s" % {
                    '_from_month': month_from, '_to_month': month_to
                })
            if year:
                filters.append(f"""DATE_PART('YEAR', hps.date_to) = {year}""")
            where_clause = f"{where_clause} {' AND '.join(filters)}"
            return where_clause
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def get_salary_report_datas(self):
        try:
            salary_report_datas = list()
            QUERY = """
            SELECT  he.id AS x_employee_id,
                he.x_social_insurance_no AS x_social_insurance_no,
                he.first_contract_date AS x_first_contract_date,
                hc.wage AS x_contract_wage,
                hc.id AS x_contract_id,
                SUM(insurances.social_insurance_amount) AS x_social_insurance_wage,
                SUM(insurances.medical_insurance_amount) AS x_medical_insurance_wage,
                SUM(insurances.unemployment_insurance_amount) AS x_unemployment_insurance_wage,
                SUM(insurances.social_insurance_by_employee_amount) AS x_social_insurance_by_employee_wage,
                SUM(insurances.medical_insurance_by_employee_amount) AS x_medical_insurance_by_employee_wage,
                SUM(insurances.unemployment_insurance_by_employee_amount) AS x_unemployment_insurance_by_employee_wage,
                SUM(insurances.social_insurance_by_company_amount) AS x_social_insurance_by_company_wage,
                SUM(insurances.medical_insurance_by_company_amount) AS x_medical_insurance_by_company_wage,
                SUM(insurances.unemployment_insurance_by_company_amount) AS x_unemployment_insurance_by_company_wage
            FROM    hr_payslip AS hps
            LEFT JOIN   hr_employee AS he ON hps.employee_id = he.id
            LEFT JOIN   hr_contract AS hc ON he.id = hc.employee_id
            LEFT JOIN   (
                SELECT  hpsl.slip_id AS slip_id,
                        CASE
                            WHEN hpsl.code = 'BHXH' THEN hpsl.total ELSE 0
                        END AS social_insurance_amount,
                        CASE
                            WHEN hpsl.code = 'BHYT' THEN hpsl.total ELSE 0
                        END AS medical_insurance_amount,
                        CASE
                            WHEN hpsl.code = 'BHTN' THEN hpsl.total ELSE 0
                        END AS unemployment_insurance_amount,
                        CASE
                            WHEN hpsl.code = 'BHXHNLD' THEN hpsl.total ELSE 0
                        END AS social_insurance_by_employee_amount,
                        CASE
                            WHEN hpsl.code = 'BHYTNLD' THEN hpsl.total ELSE 0
                        END AS medical_insurance_by_employee_amount,
                        CASE
                            WHEN hpsl.code = 'BHTNNLD' THEN hpsl.total ELSE 0
                        END AS unemployment_insurance_by_employee_amount,
                        CASE
                            WHEN hpsl.code = 'BHXHCT' THEN hpsl.total ELSE 0
                        END AS social_insurance_by_company_amount,
                        CASE
                            WHEN hpsl.code = 'BHYTCT' THEN hpsl.total ELSE 0
                        END AS medical_insurance_by_company_amount,
                        CASE
                            WHEN hpsl.code = 'BHTNCT' THEN hpsl.total ELSE 0
                        END AS unemployment_insurance_by_company_amount,
                        CASE
                            WHEN hpsl.code = 'PT' THEN hpsl.total ELSE 0
                        END AS cd
                FROM    hr_payslip_line AS hpsl
            ) AS insurances ON insurances.slip_id = hps.id
            {WHERE}
            GROUP BY he.id, he.x_social_insurance_no, he.first_contract_date, hc.wage, hc.id
            ;
            """
            filter_clause = self._generate_query_filter(
                month_from=self.x_month_from, month_to=self.x_month_to, year=self.x_year)
            self.env.cr.execute(QUERY.format(WHERE=filter_clause))
            datas = self.env.cr.dictfetchall()
            for data in datas:
                salary_report_datas.append(data)
            return salary_report_datas
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def generate_salary_reports(self):
        try:
            self.x_salary_report_ids.unlink()
            salary_reports = list()
            salary_report_vals = self.get_salary_report_datas()
            for salary_report_val in salary_report_vals:
                salary_report_val['x_report_id'] = self.id
                salary_report = self.env['social.insurance.salary.report'].create(salary_report_val)
                salary_reports.append((4, salary_report.id))
            self.x_salary_report_ids = salary_reports
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    @classmethod
    def write_a_cell(cls, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            border = Border(
                left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]
            cell.border = border

    def generate_report_file(self):
        try:
            filename = 'social_insurance_excel_template_report.xlsx'
            existing_filepath = file_path(file_path=f"enmasys_hr_payroll/static/{filename}")
            wb = load_workbook(existing_filepath)
            ws = wb.active

            data_to_write = self.get_salary_report_datas()
            start_row = 4
            index_row = 0

            for row in data_to_write:
                line = [
                    index_row + 1,
                    self.env['hr.employee'].browse(row.get("x_employee_id")).name,
                    row["x_social_insurance_no"],
                    row.get('x_first_contract_date').strftime("%d/%m/%Y") if row.get('x_first_contract_date') else "",
                    row["x_contract_wage"],
                    row.get('seniority_allowances', 0),
                    row["x_social_insurance_wage"],
                    row["x_medical_insurance_wage"],
                    row["x_unemployment_insurance_wage"],
                    row["x_social_insurance_by_employee_wage"],
                    row["x_medical_insurance_by_employee_wage"],
                    row["x_unemployment_insurance_by_employee_wage"],
                    row["x_social_insurance_by_company_wage"],
                    row["x_medical_insurance_by_company_wage"],
                    row["x_unemployment_insurance_by_company_wage"],
                ]
                self.write_a_cell(ws, line, start_row, index_row)
                index_row += 1
            fp = io.BytesIO()
            wb.save(fp)
            # create new attachment
            new_report_file = self.env['ir.attachment'].sudo().create({
                'name': filename,
                'datas': base64.encodebytes(fp.getvalue()),
                'res_model': self._name,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            # download new created attachment
            return {
                "type": "ir.actions.act_url",
                "name": filename,
                "url": '/web/content/' + str(new_report_file.id) + '?download=true',
                "target": "new",
            }
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)
