import logging
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment
import io
import base64

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from odoo.tools.misc import file_path

_logger = logging.getLogger(__name__)


class OverTimeReport(models.TransientModel):
    _name = "over.time.report"
    _description = "REPORT: OVER-TIME"
    _rec_name = 'name'


    # new
    x_department_ids = fields.Many2many(
        comodel_name="hr.department", string="Departments",
        relation="over_time_report_hr_department_rel", column1="report_id", column2="department_id")
    x_employee_ids = fields.Many2many(
        comodel_name="hr.employee", string="Employees",
        relation="over_time_report_hr_employee_rel", column1="report_id", column2="employee_id")
    x_report_year = fields.Integer(string="Year", required=True)
    x_report_month = fields.Selection(
        selection=[(str(month), f"0{month}") if month < 10 else (str(month), f"{month}") for month in range(1, 13)],
        string="Month")
    name = fields.Char(string="Tên", default='BÁO CÁO TĂNG CA')


    @classmethod
    def _handle_exception(cls, exception, raise_if_something_wrong=True):
        _logger.exception(msg=exception)
        exception = ValidationError(exception) if isinstance(exception, ValidationError) else UserError(exception)
        raise exception if raise_if_something_wrong else None

    display_name = fields.Char(string="Display name", default=_("Overtime Report"))

    @api.model
    def default_get(self, fields):
        try:
            defaults = super(OverTimeReport, self).default_get(fields)
            now = datetime.now()
            defaults['x_report_year'] = now.year
            defaults['x_report_month'] = str(now.month)
            return defaults
        except Exception as e:
            _logger.exception(msg=e)
            raise UserError(e)

    x_over_time_report_summary_line_ids = fields.One2many(
        comodel_name="over.time.report.summary.line", string="Summary lines", inverse_name="x_over_time_report_id")

    @classmethod
    def _generate_query_filter(cls, departments=None, employees=None, year=None, month=None):
        try:
            where_clause = str()
            if not departments and not employees and not year and not month:
                return where_clause
            where_clause = "WHERE"
            default_filters = [
                "ot.state = 'validate'"
            ]
            filters = default_filters
            if departments:
                filters.append("hd.id IN %(department_ids)s" % {
                    'department_ids': '(' + ', '.join(map(str, departments.ids)) + ')'
                })
            if employees:
                filters.append("he.id IN %(employee_ids)s" % {
                    'employee_ids': '(' + ', '.join(map(str, employees.ids)) + ')'
                })
            if year:
                filters.append(f"""date_part('year', otl.x_working_date) = {year}""")
            if month:
                filters.append(f"""date_part('month', otl.x_working_date) = {month}""")
            where_clause = f"{where_clause} {' AND '.join(filters)}"
            return where_clause
        except Exception as e:
            _logger.exception(msg=e)

    def _get_over_time_report_detail_line_datas(self):
        try:
            over_time_report_detail_line_datas = list()
            QUERY = """
                SELECT  otl.x_employee_id as x_employee_id,
                        hwet.id AS x_work_entry_type_id,
                        otl.x_working_date AS x_over_time_request_date,
                        otl.x_approval_hours AS x_over_time_duration,
                        otl.id AS x_over_time_line_id
                FROM    hr_overtime_interval as otl
                INNER JOIN	hr_overtime AS ot ON ot.id = otl.x_overtime_id 
                INNER JOIN	hr_employee AS he ON he.id = otl.x_employee_id 
                INNER JOIN	hr_department AS hd ON hd.id = he.department_id
                INNER JOIN	hr_work_entry_type AS hwet ON hwet.id = otl.x_working_type_id
                {WHERE};
                """
            filter_clause = self._generate_query_filter(
                departments=self.x_department_ids, employees=self.x_employee_ids,
                year=self.x_report_year, month=self.x_report_month)
            self.env.cr.execute(QUERY.format(WHERE=filter_clause))
            datas = self.env.cr.dictfetchall()
            for data in datas:
                over_time_report_detail_line_datas.append(data)
            return over_time_report_detail_line_datas
        except Exception as e:
            self._handle_exception(exception=e)

    def generate_over_time_report_summary_line(self):
        try:
            self.x_over_time_report_summary_line_ids = [(5, 0, 0)]
            report_detail_line_vals = self._get_over_time_report_detail_line_datas()
            report_detail_line_employee_ids = list(set([data.get('x_employee_id') for data in report_detail_line_vals]))
            report_summary_line_vals = {}
            for report_detail_line_employee_id in report_detail_line_employee_ids:
                report_summary_line_vals[report_detail_line_employee_id] = list(filter(
                    lambda val: val.get('x_employee_id') == report_detail_line_employee_id, report_detail_line_vals))
            report_summary_line_ids = []
            for employee_id, detail_vals in report_summary_line_vals.items():
                new_report_detail_lines = self.env['over.time.report.detail.line'].create(detail_vals)
                new_report_summary_line = self.env['over.time.report.summary.line'].create({
                    'x_over_time_report_id': self.id,
                    'x_employee_id': employee_id,
                    'x_over_time_report_detail_line_ids': new_report_detail_lines.ids
                })
                report_summary_line_ids.append((4, new_report_summary_line.id))
            self.x_over_time_report_summary_line_ids = report_summary_line_ids
        except Exception as e:
            self._handle_exception(exception=e)

    @classmethod
    def _make_cell_fit_content(cls, worksheet, cell, width_rate):
        cell_content_width = len(str(cell.value))
        current_column_width = worksheet.column_dimensions[cell.column_letter].width or 10
        new_column_width = max(current_column_width, cell_content_width) * width_rate
        worksheet.column_dimensions[cell.column_letter].width = new_column_width

    @classmethod
    def _set_report_file_main_table_content(
            cls, worksheet, main_table_lines, first_row_position=5, except_column_indexes=None):
        if not main_table_lines:
            return
        _SIDE = Side(border_style='thin', color="FF000000")
        _BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
        _ALIGNMENT = Alignment(horizontal='center')
        main_table_lines_content = cls._from_lines_generate_report_full_column_main_lines_content(
            lines=main_table_lines, except_column_indexes=except_column_indexes)
        for main_table_line_content in main_table_lines_content:
            for column_idx, line_content in main_table_line_content.items():
                line_content_cell = worksheet.cell(row=first_row_position, column=column_idx)
                line_content_cell.value = line_content
                line_content_cell.border = _BORDER
                if column_idx == 1:
                    line_content_cell.alignment = _ALIGNMENT
                cls._make_cell_fit_content(worksheet=worksheet, cell=line_content_cell, width_rate=1.0)
            first_row_position += 1

    @classmethod
    def _set_report_file_footer_content(
            cls, worksheet, summarized_lines, first_row_position=5, row_step=2, except_column_indexes=None):
        try:
            detail_lines = summarized_lines.x_over_time_report_detail_line_ids
            if not detail_lines:
                return
            work_entry_types = list(set([
                detail_line.x_work_entry_type_id for detail_line in detail_lines
            ]))
            quantity_employee_over_time_label = _("Total Employees")
            total_over_time_interval = _("Total OVER-TIME Intervals")
            work_entry_types_total = {
                work_entry_type: (
                    f"{quantity_employee_over_time_label}: {len(list(set([detail_line.x_employee_id for detail_line in detail_lines.filtered(lambda dl: dl.x_work_entry_type_id == work_entry_type)] or [])))}",
                    f"{total_over_time_interval}: {sum([detail_line.x_over_time_duration for detail_line in detail_lines.filtered(lambda dl: dl.x_work_entry_type_id == work_entry_type)] or [])}"
                )
                for work_entry_type in work_entry_types
            }
            footer_row_start_at_position = first_row_position + len(summarized_lines) + row_step
            footer_end_col_idx = 5 if except_column_indexes else 7
            footer_col_start_at_position = (
                # (row_idx, col_idx)
                len(summarized_lines), footer_end_col_idx - 2
            )
            for work_entry_type, total in work_entry_types_total.items():
                for col_idx in range(footer_col_start_at_position[1], footer_end_col_idx + 1):
                    footer_cell = worksheet.cell(row=footer_row_start_at_position, column=col_idx)
                    if col_idx == footer_col_start_at_position[1]:
                        footer_cell.value = work_entry_type.name
                        cls._make_cell_fit_content(worksheet=worksheet, cell=footer_cell, width_rate=1.0)
                        continue
                    footer_cell.value = total[1] if col_idx == footer_end_col_idx else total[0]
                    cls._make_cell_fit_content(worksheet=worksheet, cell=footer_cell, width_rate=1.0)
                footer_row_start_at_position += 1
        except Exception as e:
            cls._handle_exception(exception=e)

    @classmethod
    def _from_lines_generate_report_full_column_main_lines_content(cls, lines=None, except_column_indexes=None):
        try:
            lines_content = list()
            if not lines:
                return
            for line in lines:
                line_content = {
                    # column_index : content
                    # -> column_index: position of column in report template file
                    1: len(lines_content) + 1,
                    2: line.x_employee_id.sudo().display_name or str(),
                    3: line.x_employee_position_id.sudo().name or str(),
                    4: line.x_employee_department_id.sudo().name or str(),
                    5: line.x_total_allocated_hours,
                    6: line.x_total_validated_hours,
                    7: line.x_total_remaining_hours
                } if not except_column_indexes else {
                    # column_index : content
                    # -> column_index: position of column in report template file
                    1: len(lines_content) + 1,
                    2: line.x_employee_id.sudo().display_name or str(),
                    3: line.x_employee_position_id.sudo().name or str(),
                    4: line.x_employee_department_id.sudo().name or str(),
                    5: line.x_total_validated_hours,
                }
                lines_content.append(line_content)
            return lines_content
        except Exception as e:
            cls._handle_exception(exception=e)

    def generate_report_file(self):
        try:
            # prepare values
            summarized_lines = self.x_over_time_report_summary_line_ids
            if not summarized_lines:
                return
            report_name_suffix = f"{f'{self.x_report_month}/' if self.x_report_month else str()}{self.x_report_year}"
            report_file_name = f"BÁO CÁO TĂNG CA {report_name_suffix}"
            report_file_template_name = 'over_time_report_miss_allocated_hours_and_remaining_hours_columns.xlsx' if self.x_report_month else 'over_time_report_full_column.xlsx'
            report_file_path = file_path(file_path=f"enmasys_overtime/static/{report_file_template_name}")
            workbook = load_workbook(report_file_path)
            worksheet = workbook.active

            # write title
            report_file_title_cell = worksheet.cell(row=1, column=1)
            report_file_title_cell.value = report_file_name
            # write table content
            except_column_indexes = [5, 7] if self.x_report_month else list()
            self._set_report_file_main_table_content(
                worksheet=worksheet, main_table_lines=summarized_lines, except_column_indexes=except_column_indexes)
            # write footer
            self._set_report_file_footer_content(
                worksheet=worksheet, summarized_lines=summarized_lines, except_column_indexes=except_column_indexes, )
            # save report file data
            fp = io.BytesIO()
            workbook.save(fp)
            # create new attachment
            new_report_file = self.env['ir.attachment'].sudo().create({
                'name': report_file_name,
                'datas': base64.encodebytes(fp.getvalue()),
                'res_model': self._name,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            # download new created attachment
            return {
                "type": "ir.actions.act_url",
                "name": report_file_name,
                "url": '/web/content/' + str(new_report_file.id) + '?download=true',
                "target": "new",
            }
        except Exception as e:
            self._handle_exception(exception=e)

    x_work_entry_types_group_display_html = fields.Html(
        string="Work Entry Types display", compute="_compute_x_work_entry_types_group_display_html")

    @api.depends('x_over_time_report_summary_line_ids')
    def _compute_x_work_entry_types_group_display_html(self):
        try:
            for report in self:
                detail_lines = report.x_over_time_report_summary_line_ids.x_over_time_report_detail_line_ids
                if not detail_lines:
                    report.x_work_entry_types_group_display_html = str()
                    continue
                work_entry_types = list(set([
                    detail_line.x_work_entry_type_id for detail_line in detail_lines
                ]))
                work_entry_type_total = {
                    work_entry_type: (
                        len(list(set([detail_line.x_employee_id for detail_line in detail_lines.filtered(
                            lambda dl: dl.x_work_entry_type_id == work_entry_type)])) or []),
                        sum([detail_line.x_over_time_duration for detail_line in detail_lines.filtered(
                            lambda dl: dl.x_work_entry_type_id == work_entry_type)] or [])
                    )
                    for work_entry_type in work_entry_types
                }
                report.x_work_entry_types_group_display_html = self._generate_work_entry_type_total_html(
                    datas=work_entry_type_total)
        except Exception as e:
            self._handle_exception(exception=e)

    @classmethod
    def _generate_work_entry_type_total_html(cls, datas):
        try:
            table = """
                    <table style="white-space: nowrap; color:#000000; font-family: sans-serif; font-size: 13px; border-collapse:collapse;">
                        {ROWS}
                    </table>
                    """
            tds = []
            quantity_employee_over_time_label = _("Total Employees")
            total_over_time_interval = _("Total OVER-TIME Intervals")
            for work_entry_type, totals in datas.items():
                tds.extend([
                    "<tr>",
                    f"""<td style="text-align: left;">{work_entry_type.name}</td>""",
                    f"""<td style="text-align: left; padding-right: 10px; padding-left: 10px;">
                            {quantity_employee_over_time_label}: {totals[0]}
                        </td>""",
                    f"""<td style="text-align: left; padding-right: 10px; padding-left: 10px;">
                            {total_over_time_interval}: {totals[1]}
                        </td>""",
                    "</tr>"
                ])
            return table.format(ROWS='\n'.join(tds))
        except Exception as e:
            cls._handle_exception(exception=e)
