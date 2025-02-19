import logging
import io
import base64
from openpyxl.styles import numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from odoo.tools.misc import file_path
from openpyxl.reader.excel import load_workbook

_logger = logging.getLogger(__name__)


class HrHolidayReport(models.TransientModel):
    _name = "hr.holiday.report"
    _description = "Holiday Report"

    display_name = fields.Char(default="Báo cáo nghỉ phép", string="Display name")

    # filters
    x_department_ids = fields.Many2many(
        comodel_name="hr.department", string="Departments",
        relation="holiday_report_department_rel", column1="report_id", column2="department_id")
    x_employee_ids = fields.Many2many(
        comodel_name="hr.employee", string="Employees",
        relation="holiday_report_employee_rel", column1="report_id", column2="employee_id")
    x_year_ref = fields.Integer(string="Ref year", required=True)

    @api.model
    def default_get(self, default_fields):
        try:
            defaults = super(HrHolidayReport, self).default_get(default_fields)
            defaults['x_year_ref'] = fields.Date.today().year
            return defaults
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    x_data_ids = fields.One2many(
        comodel_name="hr.holiday.data.report", string="Report datas", inverse_name="x_report_id")

    @classmethod
    def query_data_report(cls):
        try:
            query = """
                WITH leave_allocation AS (
                    SELECT he.id AS employee_code,
                           he.name AS employee_name,
                           he.job_id AS hr_job_id,
                           he.department_id AS hr_department_id,
                           he.id AS employee_id,
                           CASE WHEN    extract(year FROM hlt.x_validity_start_date) = {year} 
                           THEN sum(hla.number_of_days) END AS quantity_leave_allocation,
                           CASE WHEN    extract(year FROM hlt.x_validity_start_date) = {last_year} 
                           THEN sum(hla.number_of_days) END AS quantity_leave_allocation_last_year
                    FROM hr_leave_allocation hla 
                    INNER JOIN hr_employee he ON he.id = hla.employee_id
                    INNER JOIN hr_leave_type hlt ON hlt.id = hla.holiday_status_id
                    INNER JOIN hr_work_entry_type hwet ON hwet.id = hlt.work_entry_type_id
                    WHERE   hla.state = 'validate'
                            AND hwet.code = 'P'
                            {department_condition}
                            {employees_condition}  
                    GROUP BY he.id, he.job_id, he.department_id, extract(year FROM hlt.x_validity_start_date)
                ),
                leave AS (
                    SELECT he.id AS employee_id,
                           CASE WHEN    extract(year FROM hl.request_date_from) = {year} 
                                        AND extract(year FROM hlt.x_validity_start_date) = {year} 
                           THEN hl.number_of_days END AS quantity_day_off,
                           CASE WHEN    extract(year FROM hl.request_date_from) = {year} 
                                        AND extract(year FROM hlt.x_validity_start_date) = {year} 
                           THEN hl.number_of_days END AS quantity_leave_used,
                           CASE WHEN    extract(year FROM hl.request_date_from) = {year}
                                        AND extract(year FROM hlt.x_validity_start_date) = {last_year}
                           THEN hl.number_of_days END AS quantity_leave_last_year_used_in_current_year,
                           CASE WHEN    extract(year FROM hl.request_date_from) = {last_year}
                                        AND extract(year FROM hlt.x_validity_start_date) = {last_year}
                           THEN hl.number_of_days END AS quantity_leave_used_last_year
                    FROM hr_leave hl
                    INNER JOIN hr_employee he ON he.id = hl.employee_id
                    INNER JOIN hr_leave_type hlt ON hlt.id = hl.holiday_status_id
                    INNER JOIN hr_work_entry_type hwet ON hwet.id = hlt.work_entry_type_id
                    WHERE   hl.state = 'validate'
                            AND hlt.time_type = 'leave'
                            AND hwet.code = 'P'
                ),
                leave_sum AS (
                    SELECT employee_id,
                           sum(quantity_day_off) AS quantity_day_off,
                           sum(quantity_leave_used) AS quantity_leave_used,
                           sum(quantity_leave_last_year_used_in_current_year) AS quantity_leave_last_year_used_in_current_year,
                           sum(quantity_leave_used_last_year) AS quantity_leave_used_last_year
                    FROM leave
                    GROUP BY employee_id
                ),
                final_result AS (
                    SELECT la.employee_code,
                           la.hr_job_id,
                           la.hr_department_id,
                           la.employee_id,
                           la.quantity_leave_allocation,
                           la.quantity_leave_allocation_last_year,
                           ls.quantity_day_off,
                           ls.quantity_leave_used,
                           CASE WHEN ls.quantity_leave_used IS NULL THEN la.quantity_leave_allocation
                                ELSE la.quantity_leave_allocation - ls.quantity_leave_used END AS quantity_remaining_leave,
                           ls.quantity_leave_last_year_used_in_current_year,
                           CASE WHEN ls.quantity_leave_used_last_year IS NULL THEN la.quantity_leave_allocation_last_year
                                ELSE la.quantity_leave_allocation_last_year - ls.quantity_leave_used_last_year END AS quantity_remaining_leave_last_year
                    FROM leave_allocation la
                    LEFT JOIN leave_sum ls ON ls.employee_id = la.employee_id
                )
                SELECT  employee_code,
                        employee_id,
                        hr_job_id,
                        hr_department_id,
                        SUM(quantity_leave_allocation) AS quantity_leave_allocation,
                        SUM(quantity_leave_allocation_last_year) AS quantity_leave_allocation_last_year,
                        quantity_day_off AS quantity_day_off,
                        quantity_leave_used AS quantity_leave_used,
                        SUM(quantity_remaining_leave) AS quantity_remaining_leave,
                        quantity_leave_last_year_used_in_current_year AS quantity_leave_last_year_used_in_current_year,
                        SUM(quantity_remaining_leave_last_year) AS quantity_remaining_leave_last_year
                FROM final_result
                GROUP BY    employee_code, employee_id, hr_job_id, hr_department_id, 
                            quantity_leave_last_year_used_in_current_year,
                            quantity_leave_used,
                            quantity_day_off
                ORDER BY employee_code
            """
            return query
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def get_data(self):
        try:
            list_employee = []

            for idd in self.x_employee_ids.ids:
                list_employee.append(idd)

            employee_condition = f" AND he.id IN ({','.join([str(idd) for idd in list_employee])})" if list_employee else str()
            department_condition = f" AND he.department_id IN ({','.join([str(idd) for idd in self.x_department_ids.ids])})" if self.x_department_ids else str()
            year = self.x_year_ref
            last_year = year - 1
            query = self.query_data_report()
            query = query.format(
                employees_condition=employee_condition, department_condition=department_condition,
                year=year, last_year=last_year)
            self.env.cr.execute(query)
            res = self.env.cr.dictfetchall()
            return res
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def open_report_view(self):
        try:
            self.x_data_ids.unlink()
            line_ids = self.get_data()
            values = []
            for line in line_ids:
                vals = {
                    'x_employee_id': line.get('employee_id'),
                    'x_remaining_holiday_last_year': line.get('quantity_remaining_leave_last_year'),
                    'x_used_of_remaining_holiday_last_year': line.get(
                        'quantity_leave_last_year_used_in_current_year'),
                    'x_allocated_intervals': line.get('quantity_leave_allocation'),
                    'x_validated_holiday_requests': line.get('quantity_day_off'),
                    'x_used_leaves': line.get('quantity_leave_used'),
                    'x_remaining_leave': line.get('quantity_remaining_leave'),
                    'x_report_id': self.id
                }
                values.append((0, 0, vals))
            self.x_data_ids = values
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    @classmethod
    def write_a_cell(cls, ws, row_data, start_row, index_row):
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=start_row + index_row, column=col)
            cell.value = row_data[col - 1]

    @classmethod
    def format_number_excel(cls, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row)]
            if currentCell.value:
                if currentCell.value % 1 == 0:
                    currentCell.number_format = '#,###'
                else:
                    currentCell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    @classmethod
    def add_border_excel(cls, ws, index_row, column_quantity):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for column in range(column_quantity):
            ws.cell(row=index_row, column=column + 1).border = thin_border

    @classmethod
    def text_align_center_excel(cls, ws, index_row, column_list):
        for column in column_list:
            currentCell = ws[str(column) + str(index_row + 1)]
            currentCell.alignment = Alignment(horizontal='center')

    def open_report_excel(self):
        try:
            filename = 'report_on_leave.xlsx'
            existing_filepath = file_path(file_path=f"enmasys_overtime/static/{filename}")
            wb = load_workbook(existing_filepath)
            ws = wb.active

            data_to_write = self.get_data()
            start_row = 2
            index_row = 0
            count = 0

            for row in data_to_write:
                count += 1
                line = [
                    count,
                    self.env['hr.employee'].sudo().browse(row['employee_id']).name,
                    self.env['hr.job'].sudo().browse(row["hr_job_id"]).name,
                    self.env['hr.department'].sudo().browse(row["hr_department_id"]).name,
                    row["quantity_remaining_leave_last_year"],
                    row["quantity_leave_last_year_used_in_current_year"],
                    row["quantity_leave_allocation"],
                    row["quantity_day_off"],
                    row["quantity_leave_used"],
                    row["quantity_remaining_leave"],
                ]
                self.write_a_cell(ws, line, start_row, index_row)
                index_row += 1
                self.add_border_excel(ws, index_row + 1, column_quantity=11)
                self.text_align_center_excel(ws, index_row, ['A'])
                self.format_number_excel(ws, index_row + 1, ['F', 'G', 'H', 'I', 'J', 'K'])

            fp = io.BytesIO()
            wb.save(fp)
            # create new attachment
            new_report_file = self.env['ir.attachment'].sudo().create({
                'name': filename,
                'datas': base64.encodebytes(fp.getvalue()),
                'res_model': self._name,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            # download new created attachment
            return {
                "type": "ir.actions.act_url",
                "name": filename,
                "url": '/web/content/' + str(new_report_file.id) + '?download=true',
                "target": "new",
            }
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)